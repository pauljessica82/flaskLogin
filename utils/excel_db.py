import openpyxl as xl


class ExcelDatabase:

    def __init__(self, file_name):
        workbook = xl.load_workbook(file_name)
        sheet = workbook['User Info']
        db_usernames = []
        db_passwords = []
        for col in sheet['E']:
            if col.value is not None:
                db_usernames.append(col.value)
                list(set(db_usernames))
        db_usernames.remove('Username')
        for col in sheet['F']:
            if col.value is not None:
                db_passwords.append(col.value)
                list(set(db_passwords))
        db_passwords.remove('Password')
        self.usernames_passwords = dict(zip(db_usernames, db_passwords))

    def invalid_credentials(self, username, password):
        return self.usernames_passwords.get(username) != password

    def add_user(self, firstname, lastname, email, phone, username, password):
        wb = xl.load_workbook('flask login.xlsx')
        sheet = wb['User Info']
        sheet.append([firstname, lastname, email, phone, username, password])
        wb.save('flask login.xlsx')

    def create_message(self, title, msg):
        wb = xl.load_workbook('flask login.xlsx')
        sheet = wb['Messages']
        sheet.append([title, msg])
        wb.save('flask login.xlsx')




